# excel_manager module
# Module holds the class => ExcelManager - manages the Excel File Interface
# Class responsible for all the excel file interface, including file name search and data pull
#
from glob import glob
from openpyxl import load_workbook


class ExcelManager(object):
    def __init__(self):
        self.watcher_columns = ['F', 'G', 'H', 'I', 'J']
        self.account_data = dict()
        self.account_file_name = ""

    # Open workbook and search for row number that corresponds to the ticket pid
    #
    def pid_row_search(self, pid, path, media_partner):
        self.account_file_name = self.get_file_name('{}/*.xlsx'.format(path))
        wb = load_workbook(filename=self.account_file_name, data_only=True)
        sheet = wb['CPG']
        for row in range(1, sheet.max_row+1):
            if sheet['B{}'.format(row)].value == pid and sheet['D{}'.format(row)].value == media_partner:
                return row

    # Open workbook and search for row number that corresponds to the ticket advertiser
    #
    def advertiser_row_search(self, advertiser, path, media_partner):
        self.account_file_name = self.get_file_name('{}/*.xlsx'.format(path))
        wb = load_workbook(filename=self.account_file_name, data_only=True)
        sheet = wb['CPG']
        for row in range(1, sheet.max_row+1):
            if sheet['A{}'.format(row)].value == advertiser and sheet['D{}'.format(row)].value == media_partner:
                return row

    # With the matched row number as input, read and save account data for email population
    #
    def excel_read(self, row):
        wb = load_workbook(filename=self.account_file_name)
        sheet = wb['CPG']

        # create a list of watchers
        watchers = list()
        for column in self.watcher_columns:
            if sheet['{}{}'.format(column, row)].value is not None:
                watchers.append(sheet['{}{}'.format(column, row)].value)

        # create a dictionary of values from excel data
        self.account_data = {
            "account_type":         sheet['C{}'.format(row)].value,
            "media_partner":        sheet['D{}'.format(row)].value,
            "solutions_mgr":        sheet['F{}'.format(row)].value,
            "solutions_manager":    sheet['G{}'.format(row)].value,
            "client_analytics":     sheet['J{}'.format(row)].value,
            "watchers":             watchers
        }
        return self.account_data

    # Search specified folder for what should be the only file, get name and return
    #
    @staticmethod
    def get_file_name(path):
        for file_name in glob(path):
            return file_name
